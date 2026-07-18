import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel, scorer


BAD_DOMAIN_KEYWORDS = {
    "hotel",
    "gazete",
    "gazetesi",
    "gov",
    "belediye",
    "haber",
    "news",
    "chef",
    "cert",
    "bilisim",
    "savunma",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Old/new karsilastirma dosyasina karar onerisi ekler.")
    parser.add_argument("--input", type=Path, default=config.OUTPUT_DIR / "old_vs_brightdata_decisions.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "old_vs_brightdata_suggested.xlsx")
    return parser.parse_args()


def _load(path: Path) -> tuple[list[str], list[dict]]:
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    data = [
        {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        for row in rows[1:]
    ]
    return headers, data


def _domain(url: str) -> str:
    return scorer.normalize_domain(url)


def _token_hits(company: str, url: str) -> tuple[int, int]:
    tokens = scorer.distinctive_tokens(company)
    compact = scorer.compact_domain_core(url)
    return sum(1 for token in tokens if token in compact), len(tokens)


def _has_bad_keyword(url: str) -> bool:
    domain = _domain(url)
    return any(keyword in domain for keyword in BAD_DOMAIN_KEYWORDS) or scorer.is_excluded_domain(domain)


def _contact_count(row: dict, prefix: str) -> int:
    return int(bool(row.get(f"{prefix}_email"))) + int(bool(row.get(f"{prefix}_phone")))


def _tr_bonus(url: str) -> int:
    domain = _domain(url)
    return int(domain.endswith(".com.tr") or domain.endswith(".tr"))


def _score(row: dict, prefix: str) -> int:
    try:
        return int(row.get(f"{prefix}_score") or 0)
    except ValueError:
        return 0


def suggest(row: dict) -> tuple[str, str]:
    company = str(row.get("company", ""))
    old_url = str(row.get("old_website", ""))
    new_url = str(row.get("new_website", ""))
    old_hits, token_count = _token_hits(company, old_url)
    new_hits, _ = _token_hits(company, new_url)
    old_bad = _has_bad_keyword(old_url)
    new_bad = _has_bad_keyword(new_url)
    old_contacts = _contact_count(row, "old")
    new_contacts = _contact_count(row, "new")
    old_score = _score(row, "old")
    new_score = _score(row, "new")
    old_strength = old_hits * 3 + old_contacts * 2 + _tr_bonus(old_url)
    new_strength = new_hits * 3 + new_contacts * 2 + _tr_bonus(new_url)

    if not old_url and new_url and not new_bad:
        return "new", "old_empty_new_found"
    if old_bad and not new_bad:
        return "new", "old_bad_domain"
    if new_bad and not old_bad:
        return "old", "new_bad_domain"
    if old_hits == token_count and new_hits < token_count:
        return "old", f"old_full_token_match:{old_hits}/{token_count}"
    if new_hits == token_count and old_hits < token_count:
        return "new", f"new_full_token_match:{new_hits}/{token_count}"
    if new_contacts > old_contacts and new_hits >= old_hits and not new_bad:
        return "new", "new_more_contact_evidence"
    if old_contacts > new_contacts and old_hits >= new_hits and not old_bad:
        return "old", "old_more_contact_evidence"
    if old_url.rstrip("/") == new_url.rstrip("/"):
        return "old", "same_domain_keep_old"
    if abs(new_score - old_score) >= 25:
        if new_score > old_score and not new_bad:
            return "new", "new_score_much_higher"
        if old_score > new_score and not old_bad:
            return "old", "old_score_much_higher"
    if abs(new_strength - old_strength) >= 3:
        if new_strength > old_strength and not new_bad:
            return "new", f"new_stronger_evidence:{new_strength}>{old_strength}"
        if old_strength > new_strength and not old_bad:
            return "old", f"old_stronger_evidence:{old_strength}>{new_strength}"
    return "review", f"unclear old_hits:{old_hits}/{token_count} new_hits:{new_hits}/{token_count}"


def main() -> None:
    args = parse_args()
    headers, rows = _load(args.input)
    output_rows = []
    counts = {"old": 0, "new": 0, "review": 0}
    for row in rows:
        decision, reason = suggest(row)
        row["suggested_decision"] = decision
        row["suggestion_reason"] = reason
        counts[decision] += 1
        output_rows.append(row)

    output_headers = [header for header in headers if header != "decision"] + [
        "suggested_decision",
        "suggestion_reason",
        "decision",
    ]
    excel._write_rows(args.output, output_headers, output_rows)
    print(f"old onerisi: {counts['old']}")
    print(f"new onerisi: {counts['new']}")
    print(f"review kalan: {counts['review']}")
    print(f"Yazildi: {args.output}")


if __name__ == "__main__":
    main()
