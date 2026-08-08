"""Run the 45-company Türkiye-only golden benchmark without fair evidence."""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config
import main
from modules import scorer
from run_foodist_blind import configure_blind_run


ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "outputs" / "golden_lists_20260804" / "goldenT_combined_blind_45.xlsx"
DEFAULT_STATE = ROOT / "state" / "golden_tiers_loop"
OUTPUT_ROOT = ROOT / "output"
TIERS = ("goldenT1", "goldenT2", "goldenT3")
GOLDEN_DIR = ROOT / "outputs" / "golden_lists_20260804"


def _result_rows(path: Path) -> dict[str, dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()
    headers = [str(value or "").strip().casefold() for value in rows[0]]
    return {
        str(row[headers.index("company")] or "").strip().casefold(): {
            field: row[headers.index(field)] if field in headers else ""
            for field in ("website", "email", "phone")
        }
        for row in rows[1:]
        if row[headers.index("company")]
    }


def _coverage(output_dir: Path) -> dict[str, dict[str, int]]:
    combined = _result_rows(output_dir / "contacts.xlsx")
    combined.update(_result_rows(output_dir / "review_queue.xlsx"))
    result: dict[str, dict[str, int]] = {}
    for tier in TIERS:
        truth_book = load_workbook(
            GOLDEN_DIR / f"{tier}.xlsx", read_only=True, data_only=True,
        )
        truth_rows = list(truth_book.active.iter_rows(values_only=True))[4:19]
        truth_book.close()
        result[tier] = {field: 0 for field in ("website", "email", "phone")}
        for row in truth_rows:
            company = str(row[1] or "").strip()
            expected_website = str(row[4] or "").strip()
            actual = combined.get(company.casefold(), {})
            website_correct = bool(
                expected_website
                and actual.get("website")
                and scorer.same_registrable_domain(
                    expected_website, str(actual["website"]),
                )
            )
            result[tier]["website"] += int(website_correct)
            for field, truth_index in (("email", 5), ("phone", 6)):
                expected_field = str(row[truth_index] or "").strip()
                actual_field = str(actual.get(field, "") or "").strip()
                result[tier][field] += int(
                    bool(website_correct and expected_field and actual_field)
                )
    result["total"] = {
        field: sum(result[tier][field] for tier in TIERS)
        for field in ("website", "email", "phone")
    }
    return result


def main_cli() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--state-dir", type=Path, default=DEFAULT_STATE)
    parser.add_argument("--iteration", type=int, required=True)
    args = parser.parse_args()
    if not INPUT.exists():
        raise SystemExit(f"Missing benchmark input: {INPUT}")

    configure_blind_run(args.state_dir)
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = OUTPUT_ROOT / f"golden_tiers_iter{args.iteration}_{suffix}"
    print(f"Golden tiers output: {output_dir}")
    print(f"Golden tiers state: {args.state_dir}")
    print(main.run(INPUT, output_dir))
    coverage = _coverage(output_dir)
    print("GOLDEN_TIER_COVERAGE")
    for tier in (*TIERS, "total"):
        values = coverage[tier]
        denominator = 45 if tier == "total" else 15
        print(
            f"{tier}: website={values['website']}/{denominator}; "
            f"email={values['email']}/{denominator}; phone={values['phone']}/{denominator}"
        )


if __name__ == "__main__":
    main_cli()
