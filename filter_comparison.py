import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Karsilastirma dosyasindan karar gerektiren satirlari suzer.")
    parser.add_argument("--input", type=Path, default=config.OUTPUT_DIR / "old_vs_brightdata_full.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "old_vs_brightdata_decisions.xlsx")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ws = load_workbook(args.input, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    selected = []
    for row in rows[1:]:
        item = {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        old_site = item.get("old_website", "")
        new_site = item.get("new_website", "")
        if new_site and (not old_site or old_site != new_site):
            selected.append(item)

    excel._write_rows(args.output, headers, selected)
    print(f"Karar gerektiren satir: {len(selected)}")
    print(f"Yazildi: {args.output}")


if __name__ == "__main__":
    main()
