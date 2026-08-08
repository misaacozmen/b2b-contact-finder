from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def read_companies(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    first_cell = str(rows[0][0] or "").strip().lower()
    start_index = 1 if first_cell == "company" else 0
    companies: list[str] = []
    for row in rows[start_index:]:
        value = row[0] if row else None
        if value is None:
            continue
        company = str(value).strip()
        if company:
            companies.append(company)
    return companies


def read_company_records(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    first_row = [str(value or "").strip().lower() for value in rows[0]]
    has_header = "company" in first_row
    headers = first_row if has_header else []
    start_index = 1 if has_header else 0

    def value_for(row: tuple, names: tuple[str, ...], fallback_index: int | None = None) -> str:
        for name in names:
            if name in headers:
                idx = headers.index(name)
                if idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
        if (
            not has_header
            and fallback_index is not None
            and fallback_index < len(row)
            and row[fallback_index] is not None
        ):
            return str(row[fallback_index]).strip()
        return ""

    records: list[dict] = []
    for row in rows[start_index:]:
        company = value_for(row, ("company", "firma", "firma adi", "firma adı"), 0)
        if not company:
            continue
        records.append(
            {
                "company": company,
                "website": value_for(row, ("website", "web sitesi", "websitesi", "site"), 1),
                "listed_website": value_for(
                    row, ("listed_website", "fair_website", "fuar web sitesi"), None
                ),
                "source": value_for(row, ("source", "kaynak"), None),
                "country": value_for(row, ("country", "ulke", "ülke"), None),
                "profile_url": value_for(row, ("profile_url", "profil", "profile"), None),
                "listing_url": value_for(row, ("listing_url", "liste_url", "liste url"), None),
                "listed_phone": value_for(row, ("listed_phone", "fair_phone", "fuar telefonu"), None),
                "listed_email": value_for(row, ("listed_email", "fair_email", "fuar e-posta"), None),
                "listed_address": value_for(row, ("listed_address", "fair_address", "fuar adresi"), None),
                "hall": value_for(row, ("hall", "salon"), None),
                "stand": value_for(row, ("stand", "stant"), None),
                "brands": value_for(row, ("brands", "markalar"), None),
                "representations": value_for(row, ("representations", "temsilcilikler"), None),
                "sector": value_for(row, ("sector", "sektor", "sektör", "urun grubu", "ürün grubu"), None),
                "description": value_for(row, ("description", "aciklama", "açıklama"), None),
            }
        )
    return records


def read_result_statuses(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return {}
    headers = [str(value or "").strip().casefold() for value in rows[0]]
    if "company" not in headers or "status" not in headers:
        return {}
    company_idx = headers.index("company")
    status_idx = headers.index("status")
    return {
        str(row[company_idx]).strip().casefold(): str(row[status_idx] or "").strip()
        for row in rows[1:]
        if len(row) > max(company_idx, status_idx) and row[company_idx]
    }


def _write_rows(path: Path, headers: list[str], rows: Iterable[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(path)


def write_contacts(path: Path, rows: Iterable[dict]) -> None:
    _write_rows(
        path,
        [
            "company",
            "website",
            "website_source",
            "website_status",
            "email",
            "email_source",
            "email_source_url",
            "alternative_emails",
            "alternative_email_sources",
            "email_verification",
            "email_verification_reason",
            "email_publication_status",
            "email_publication_reason",
            "phone",
            "phone_source",
            "phone_source_url",
            "phone_label",
            "alternative_phones",
            "alternative_phone_sources",
            "phone_publication_status",
            "phone_publication_reason",
            "contact_policy_version",
            "contact_status",
            "status",
            "confidence",
            "score",
            "publication_policy_version",
            "publication_policy_action",
            "publication_eligible",
            "publication_safety_score",
            "publication_risk_index",
            "publication_risk_tier",
            "publication_blockers",
            "reason",
        ],
        rows,
    )


def write_failed(path: Path, rows: Iterable[dict]) -> None:
    _write_rows(path, ["company", "status", "reason"], rows)


def write_website_candidates(path: Path, rows: Iterable[dict]) -> None:
    headers = [
        "company",
        "selected_website",
        "status",
        "confidence",
        "publication_policy_version",
        "publication_policy_action",
        "publication_eligible",
        "publication_safety_score",
        "publication_risk_index",
        "publication_risk_tier",
        "publication_blockers",
        "candidate_1_url",
        "candidate_1_score",
        "candidate_1_reason",
        "candidate_1_query",
        "candidate_1_role",
        "candidate_2_url",
        "candidate_2_score",
        "candidate_2_reason",
        "candidate_2_query",
        "candidate_2_role",
        "candidate_3_url",
        "candidate_3_score",
        "candidate_3_reason",
        "candidate_3_query",
        "candidate_3_role",
    ]
    _write_rows(path, headers, rows)


def write_company_records(path: Path, rows: Iterable[dict]) -> None:
    _write_rows(
        path,
        [
            "company", "website", "listed_website", "source", "country",
            "profile_url", "listing_url", "listed_phone", "listed_email",
            "listed_address", "hall", "stand", "brands", "representations",
            "sector", "description",
        ],
        rows,
    )
