import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel, scorer

SUSPICIOUS_DOMAIN_KEYWORDS = [
    "hotel",
    "chef",
    "cert",
    "united",
    "belediye",
    "municipality",
    "university",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Yuksek skorlu ama supheli contact satirlarini ayiklar.")
    parser.add_argument("--input", type=Path, default=config.CONTACTS_FILE, help="Denetlenecek contacts.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "suspicious_contacts.xlsx", help="Supheli cikti")
    return parser.parse_args()


def _load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        for row in rows[1:]
    ]


def _domain_hit_ratio(reason: str) -> tuple[int, int] | None:
    match = re.search(r"domain_hits:(\d+)/(\d+)", reason or "")
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _is_suspicious(row: dict) -> tuple[bool, str]:
    company = str(row.get("company", ""))
    website = str(row.get("website", ""))
    status = str(row.get("status", ""))
    reason = str(row.get("reason", ""))
    score = int(row.get("score") or 0)
    if not website:
        return False, ""
    if not status.startswith("OK") and score < config.HIGH_CONFIDENCE_SCORE:
        return False, ""

    reasons = []
    domain = scorer.normalize_domain(website)
    if scorer.is_excluded_domain(domain):
        reasons.append("excluded_domain")

    hits = _domain_hit_ratio(reason)
    if hits:
        hit_count, token_count = hits
        if token_count >= 2 and hit_count < token_count:
            reasons.append(f"partial_domain_match:{hit_count}/{token_count}")

    if any(keyword in domain for keyword in SUSPICIOUS_DOMAIN_KEYWORDS):
        reasons.append("suspicious_domain_keyword")

    if status.startswith("OK") and ("no_context_tokens" in reason or "page_identity_medium" in reason):
        if reasons:
            reasons.append("ok_without_strong_context")

    if status.startswith("OK") and not row.get("email"):
        if reasons:
            reasons.append("ok_without_email")

    return bool(reasons), "; ".join(dict.fromkeys(reasons))


def main() -> None:
    args = parse_args()
    rows = _load_rows(args.input)
    suspicious = []
    for row in rows:
        is_suspicious, audit_reason = _is_suspicious(row)
        if is_suspicious:
            item = dict(row)
            item["audit_reason"] = audit_reason
            suspicious.append(item)

    excel._write_rows(
        args.output,
        ["company", "website", "email", "phone", "status", "confidence", "score", "audit_reason", "reason"],
        suspicious,
    )
    print(f"Supheli satir: {len(suspicious)}")
    print(f"Yazildi: {args.output}")


if __name__ == "__main__":
    main()
