import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel


DEFAULT_STATUSES = {
    "REVIEW_NEEDED",
    "WEBSITE_NOT_FOUND",
    "WEBSITE_FETCH_FAILED",
    "SEARCH_FAILED",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sorunlu contact satirlarindan tekrar deneme input'u uretir.")
    parser.add_argument("--contacts", type=Path, default=config.CONTACTS_FILE, help="Okunacak contacts.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "review_retry_input.xlsx", help="Olusacak input Excel")
    parser.add_argument(
        "--status",
        action="append",
        default=[],
        help="Tekrar denenecek status. Birden fazla kez verilebilir. Varsayilan: review ve failed statusleri.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Status filtrelemeden dosyadaki tum firmalari tekrar deneme input'una alir.",
    )
    return parser.parse_args()


def load_retry_rows(path: Path, statuses: set[str], include_all: bool = False) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip().lower() for value in rows[0]]
    company_idx = headers.index("company")
    status_idx = headers.index("status")

    retry_rows = []
    seen = set()
    for row in rows[1:]:
        company = str(row[company_idx] or "").strip()
        status = str(row[status_idx] or "").strip()
        if not company or (not include_all and status not in statuses):
            continue
        key = company.casefold()
        if key in seen:
            continue
        seen.add(key)
        retry_rows.append(
            {
                "company": company,
                "website": "",
                "source": f"retry_{status.lower()}",
                "country": "",
                "profile_url": "",
            }
        )
    return retry_rows


def main() -> None:
    args = parse_args()
    statuses = {status.strip() for status in args.status if status.strip()} or DEFAULT_STATUSES
    rows = load_retry_rows(args.contacts, statuses, include_all=args.all)
    excel.write_company_records(args.output, rows)
    print(f"Tekrar denenecek firma: {len(rows)}")
    print(f"Yazildi: {args.output}")


if __name__ == "__main__":
    main()
