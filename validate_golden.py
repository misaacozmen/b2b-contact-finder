"""Compare contacts.xlsx with a human-verified golden CSV.

Golden CSV columns: company,website,email,phone
Leave an expected field empty when it is intentionally not asserted.
"""

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

from modules import scorer


def _read_xlsx(path: Path) -> dict[str, dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip().lower() for value in rows[0]]
    return {
        scorer.normalize_text(str(row[headers.index("company")] or "")): {
            header: str(row[index] or "").strip() for index, header in enumerate(headers)
        }
        for row in rows[1:]
        if row and row[headers.index("company")]
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--expected", type=Path, required=True)
    parser.add_argument("--actual", type=Path, required=True)
    args = parser.parse_args()

    actual = _read_xlsx(args.actual)
    total = matched = 0
    with args.expected.open(encoding="utf-8-sig", newline="") as handle:
        for expected in csv.DictReader(handle):
            company = expected.get("company", "").strip()
            if not company:
                continue
            total += 1
            found = actual.get(scorer.normalize_text(company), {})
            checks = [
                field for field in ("website", "email", "phone")
                if expected.get(field, "").strip()
            ]
            failures = [field for field in checks if expected[field].strip().lower() not in found.get(field, "").lower()]
            if failures:
                print(f"FAIL\t{company}\t{','.join(failures)}")
            else:
                matched += 1
                print(f"PASS\t{company}")
    print(f"Golden match: {matched}/{total} ({(matched / total * 100) if total else 0:.1f}%)")


if __name__ == "__main__":
    main()
