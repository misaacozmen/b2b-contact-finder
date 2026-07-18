import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Eski ana ciktiya onerilen old/new kararlarini uygular.")
    parser.add_argument("--base", type=Path, default=config.OUTPUT_DIR / "contacts.xlsx")
    parser.add_argument("--new", type=Path, default=config.OUTPUT_DIR / "brightdata_full_rerun" / "contacts.xlsx")
    parser.add_argument("--decisions", type=Path, default=config.OUTPUT_DIR / "old_vs_brightdata_suggested.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "final_merged_contacts.xlsx")
    parser.add_argument("--review-output", type=Path, default=config.OUTPUT_DIR / "final_merged_review_left.xlsx")
    return parser.parse_args()


def _load(path: Path) -> tuple[list[str], list[dict]]:
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    data = [
        {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        for row in rows[1:]
    ]
    return headers, data


def _key(row: dict) -> str:
    return str(row.get("company", "")).strip().casefold()


def _contacts_row(row: dict) -> dict:
    return {
        "company": row.get("company", ""),
        "website": row.get("website", ""),
        "email": row.get("email", ""),
        "phone": row.get("phone", ""),
        "status": row.get("status", ""),
        "confidence": row.get("confidence", ""),
        "score": row.get("score", ""),
        "reason": row.get("reason", ""),
    }


def main() -> None:
    args = parse_args()
    _, base_rows = _load(args.base)
    _, new_rows = _load(args.new)
    _, decision_rows = _load(args.decisions)

    new_by_key = {_key(row): row for row in new_rows}
    decision_by_key = {_key(row): row for row in decision_rows}
    final_rows = []
    review_left = []
    counts = {"old": 0, "new": 0, "review": 0, "missing": 0}

    for base_row in base_rows:
        key = _key(base_row)
        decision_row = decision_by_key.get(key)
        if not decision_row:
            final_rows.append(_contacts_row(base_row))
            counts["missing"] += 1
            continue

        decision = str(decision_row.get("suggested_decision") or "").strip().lower()
        if decision == "new" and key in new_by_key:
            merged = _contacts_row(new_by_key[key])
            merged["reason"] = f"merged_from_brightdata; {merged.get('reason', '')}".strip("; ")
            final_rows.append(merged)
            counts["new"] += 1
        elif decision == "old":
            final_rows.append(_contacts_row(base_row))
            counts["old"] += 1
        else:
            final_rows.append(_contacts_row(base_row))
            review_left.append(decision_row)
            counts["review"] += 1

    excel.write_contacts(args.output, final_rows)
    excel._write_rows(
        args.review_output,
        [
            "company",
            "old_website",
            "old_email",
            "old_phone",
            "old_status",
            "old_score",
            "new_website",
            "new_email",
            "new_phone",
            "new_status",
            "new_score",
            "suggested_decision",
            "suggestion_reason",
            "decision",
        ],
        review_left,
    )
    print(f"Eski korunan karar: {counts['old']}")
    print(f"Yeniyle degistirilen karar: {counts['new']}")
    print(f"Review kalip eski korunan: {counts['review']}")
    print(f"Karar dosyasinda olmayan, eski korunan: {counts['missing']}")
    print(f"Final yazildi: {args.output}")
    print(f"Review kalan yazildi: {args.review_output}")


if __name__ == "__main__":
    main()
