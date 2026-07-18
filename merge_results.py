import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel


OK_STATUSES = {"OK_HIGH_CONFIDENCE", "OK_MEDIUM_CONFIDENCE"}
REVIEW_STATUSES = {"REVIEW_NEEDED"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ana ciktilar ile Bright Data tekrar kosusunu birlestirir.")
    parser.add_argument("--base", type=Path, default=config.CONTACTS_FILE, help="Ana contacts.xlsx")
    parser.add_argument(
        "--retry",
        type=Path,
        default=config.OUTPUT_DIR / "brightdata_review" / "contacts.xlsx",
        help="Bright Data contacts.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=config.OUTPUT_DIR / "final_contacts.xlsx",
        help="Birlesik final contacts.xlsx",
    )
    parser.add_argument(
        "--review-output",
        type=Path,
        default=config.OUTPUT_DIR / "merge_review.xlsx",
        help="Manuel kontrol gereken Bright Data satirlari",
    )
    parser.add_argument(
        "--auto-replace-ok",
        action="store_true",
        help="Bright Data OK satirlarini otomatik final dosyasina al. Varsayilan guvenli modda hepsi review dosyasina yazilir.",
    )
    return parser.parse_args()


def _load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        item = {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        if item.get("company"):
            result.append(item)
    return result


def _company_key(row: dict) -> str:
    return str(row.get("company", "")).strip().casefold()


def merge_rows(base_rows: list[dict], retry_rows: list[dict], auto_replace_ok: bool) -> tuple[list[dict], list[dict], dict]:
    retry_by_company = {_company_key(row): row for row in retry_rows if _company_key(row)}
    final_rows = []
    review_rows = []
    stats = {
        "base_total": len(base_rows),
        "retry_total": len(retry_rows),
        "auto_replaced": 0,
        "review_candidates": 0,
        "unchanged": 0,
    }

    for base_row in base_rows:
        key = _company_key(base_row)
        retry_row = retry_by_company.get(key)
        if not retry_row:
            final_rows.append(base_row)
            stats["unchanged"] += 1
            continue

        retry_status = str(retry_row.get("status", ""))
        if retry_status in OK_STATUSES and auto_replace_ok:
            merged = dict(retry_row)
            merged["reason"] = f"brightdata_retry_replaced; {merged.get('reason', '')}".strip("; ")
            final_rows.append(merged)
            stats["auto_replaced"] += 1
        elif retry_row.get("website") and retry_status in (OK_STATUSES | REVIEW_STATUSES):
            final_rows.append(base_row)
            review_item = dict(retry_row)
            review_item["suggested_action"] = "replace" if retry_status in OK_STATUSES else "manual_check"
            review_item["base_status"] = base_row.get("status", "")
            review_item["base_website"] = base_row.get("website", "")
            review_item["base_email"] = base_row.get("email", "")
            review_item["base_phone"] = base_row.get("phone", "")
            review_rows.append(review_item)
            stats["review_candidates"] += 1
        else:
            final_rows.append(base_row)
            stats["unchanged"] += 1

    return final_rows, review_rows, stats


def main() -> None:
    args = parse_args()
    base_rows = _load_rows(args.base)
    retry_rows = _load_rows(args.retry)
    final_rows, review_rows, stats = merge_rows(base_rows, retry_rows, args.auto_replace_ok)

    excel.write_contacts(args.output, final_rows)
    excel._write_rows(
        args.review_output,
        [
            "company",
            "website",
            "email",
            "phone",
            "status",
            "confidence",
            "score",
            "reason",
            "suggested_action",
            "base_status",
            "base_website",
            "base_email",
            "base_phone",
        ],
        review_rows,
    )

    print(f"Ana toplam: {stats['base_total']}")
    print(f"Bright Data tekrar toplam: {stats['retry_total']}")
    print(f"Otomatik degistirilen OK kayit: {stats['auto_replaced']}")
    print(f"Manuel kontrol adaylari: {stats['review_candidates']}")
    print(f"Degismeyen: {stats['unchanged']}")
    print(f"Final yazildi: {args.output}")
    print(f"Review yazildi: {args.review_output}")


if __name__ == "__main__":
    main()
