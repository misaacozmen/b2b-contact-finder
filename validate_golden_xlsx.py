"""Compare a pipeline contacts.xlsx directly with the manual golden workbook."""

import argparse
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from modules import phone, scorer


FIELDS = ("website", "email", "phone")
EXPECTED_COLUMNS = {
    "website": "Expected Website",
    "email": "Expected Email",
    "phone": "Expected Phone",
}
VERIFICATION_COLUMNS = {
    "website": "Website Verified",
    "email": "Email Verified",
    "phone": "Phone Verified",
}
PRESENT_VALUES = {"yes", "present", "var", "evet"}
ABSENT_VALUES = {"no", "absent", "yok", "hayir", "hayır"}
UNKNOWN_VALUES = {"unknown", "unverified", "bilinmiyor", "bilinmiyor/unknown", "dogrulanamadi", "doğrulanamadı"}


def _verification_state(value: object) -> str:
    normalized = scorer.normalize_text(str(value or "")).strip()
    normalized_sets = {
        "present": {scorer.normalize_text(item) for item in PRESENT_VALUES},
        "absent": {scorer.normalize_text(item) for item in ABSENT_VALUES},
        "unknown": {scorer.normalize_text(item) for item in UNKNOWN_VALUES},
    }
    for state, values in normalized_sets.items():
        if normalized in values:
            return state
    return ""


def _host(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parsed = urlparse(text if "://" in text else f"https://{text}")
    return (parsed.hostname or "").lower().removeprefix("www.")


def _hosts(value: object) -> list[str]:
    hosts = []
    for part in re.split(r"[\n;,|]+", str(value or "")):
        host = _host(part)
        if host and host not in hosts:
            hosts.append(host)
    return hosts


def _emails(value: object) -> list[str]:
    return re.findall(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}", str(value or "").lower())


def _phones(value: object) -> list[str]:
    normalized = []
    for part in re.split(r"[\n;,|]+", str(value or "")):
        if not part.strip():
            continue
        variants = [part]
        range_match = re.match(r"^(.+\b\d{1,2})\s*-\s*(\d{1,2})$", part.strip())
        if range_match and len(re.sub(r"\D", "", range_match.group(1))) >= 10:
            base, suffix = range_match.groups()
            variants = [base, re.sub(r"\d{1,2}\s*$", suffix, base)]
        for variant in variants:
            value_normalized = phone.normalize_phone(variant)
            if value_normalized:
                normalized.append(value_normalized)
                continue
            digits = re.sub(r"\D", "", variant)
            if digits:
                normalized.append(f"0{digits}" if len(digits) == 10 else digits)
    return list(dict.fromkeys(normalized))


def _sheet_rows(path: Path, sheet_name: str | None = None) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
    finally:
        workbook.close()


def readiness_issues(expected_path: Path, companies: set[str] | None = None) -> list[str]:
    issues = []
    selected = {scorer.normalize_text(value) for value in companies} if companies else None
    for row in _sheet_rows(expected_path, "Manual Report"):
        company = str(row.get("Company") or "").strip()
        if not company or (selected is not None and scorer.normalize_text(company) not in selected):
            continue
        for field in FIELDS:
            verification = _verification_state(row.get(VERIFICATION_COLUMNS[field]))
            raw_expected = row.get(EXPECTED_COLUMNS[field])
            expected_values = (
                _hosts(raw_expected) if field == "website"
                else _emails(raw_expected) if field == "email"
                else _phones(raw_expected)
            )
            if not verification:
                issues.append(f"{company}: {field} Verified present/absent/unknown secilmemis")
            elif verification == "present" and not expected_values:
                issues.append(f"{company}: {field} present ama Expected alani bos")
            elif verification == "absent" and expected_values:
                issues.append(f"{company}: {field} absent ama Expected alaninda deger var")
    return issues


def evaluate(expected_path: Path, actual_path: Path, companies: set[str] | None = None) -> tuple[dict, list[str]]:
    expected_rows = _sheet_rows(expected_path, "Manual Report")
    actual_rows = _sheet_rows(actual_path)
    actual_by_company = {
        scorer.normalize_text(str(row.get("company") or "").strip()): row for row in actual_rows
    }
    metrics = {field: {"tp": 0, "fp": 0, "fn": 0} for field in FIELDS}
    complete_matches: list[str] = []
    selected = {scorer.normalize_text(value).strip() for value in companies} if companies else None

    for expected in expected_rows:
        company = str(expected.get("Company") or "").strip()
        if not company:
            continue
        if selected is not None and scorer.normalize_text(company).strip() not in selected:
            continue
        actual = actual_by_company.get(scorer.normalize_text(company), {})
        expected_values = {
            "website": _hosts(expected.get("Expected Website")),
            "email": _emails(expected.get("Expected Email")),
            "phone": _phones(expected.get("Expected Phone")),
        }
        primary_values = {
            "website": _host(actual.get("website")),
            "email": str(actual.get("email") or "").strip().lower(),
            "phone": phone.normalize_phone(str(actual.get("phone") or "")),
        }
        # A verified Golden value counts as found even when the pipeline keeps
        # it in an audit-friendly alternative column.  For a Golden "no", only
        # the published primary value is asserted; unverified alternatives do
        # not create a false positive by themselves.
        actual_values = {
            "website": [primary_values["website"]] if primary_values["website"] else [],
            "email": list(dict.fromkeys([
                *([primary_values["email"]] if primary_values["email"] else []),
                *_emails(actual.get("alternative_emails")),
            ])),
            "phone": list(dict.fromkeys([
                *([primary_values["phone"]] if primary_values["phone"] else []),
                *_phones(actual.get("alternative_phones")),
            ])),
        }
        verification = {
            field: _verification_state(expected.get(VERIFICATION_COLUMNS[field]))
            for field in FIELDS
        }
        asserted = {field: verification[field] in {"present", "absent"} for field in FIELDS}
        matches = {
            field: bool(set(actual_values[field]) & set(expected_values[field]))
            for field in FIELDS
        }
        for field in FIELDS:
            if not asserted[field]:
                continue
            expects_value = verification[field] == "present"
            field_matches = matches[field] if expects_value else not primary_values[field]
            if field_matches:
                if expects_value:
                    metrics[field]["tp"] += 1
            else:
                if primary_values[field]:
                    metrics[field]["fp"] += 1
                if expects_value and expected_values[field]:
                    metrics[field]["fn"] += 1
        if all(asserted.values()) and all(
            matches[field] if verification[field] == "present" else not primary_values[field]
            for field in FIELDS
        ):
            complete_matches.append(company)

    return metrics, complete_matches


def assertion_coverage(expected_path: Path) -> dict[str, dict[str, int]]:
    coverage = {field: {"asserted": 0, "unknown": 0, "missing": 0} for field in FIELDS}
    for row in _sheet_rows(expected_path, "Manual Report"):
        if not str(row.get("Company") or "").strip():
            continue
        for field in FIELDS:
            state = _verification_state(row.get(VERIFICATION_COLUMNS[field]))
            if state in {"present", "absent"}:
                coverage[field]["asserted"] += 1
            elif state == "unknown":
                coverage[field]["unknown"] += 1
            else:
                coverage[field]["missing"] += 1
    return coverage


def evaluate_stages(expected_path: Path, actual_path: Path, candidates_path: Path) -> dict:
    """Measure discovery, selection, publication and extraction as separate stages."""
    expected_rows = _sheet_rows(expected_path, "Manual Report")
    actual_by_company = {
        scorer.normalize_text(str(row.get("company") or "").strip()): row
        for row in _sheet_rows(actual_path)
    }
    candidates_by_company = {
        scorer.normalize_text(str(row.get("company") or "").strip()): row
        for row in _sheet_rows(candidates_path)
    }
    counts = {
        "expected_websites": 0, "website_asserted_rows": 0,
        "website_unknown_rows": 0, "candidate_top1_hits": 0, "candidate_top3_hits": 0,
        "selected_count": 0, "selected_correct": 0,
        "published_count": 0, "published_correct": 0, "abstained": 0,
        "correct_site_rows": 0, "correct_site_email_asserted": 0,
        "correct_site_phone_asserted": 0,
        "email_on_correct_site": 0, "phone_on_correct_site": 0,
    }
    total = 0
    for expected in expected_rows:
        company = str(expected.get("Company") or "").strip()
        if not company:
            continue
        total += 1
        key = scorer.normalize_text(company)
        expected_hosts = set(_hosts(expected.get("Expected Website")))
        website_state = _verification_state(expected.get(VERIFICATION_COLUMNS["website"]))
        email_state = _verification_state(expected.get(VERIFICATION_COLUMNS["email"]))
        phone_state = _verification_state(expected.get(VERIFICATION_COLUMNS["phone"]))
        # Older benchmark fixtures predate explicit verification columns. Keep
        # them usable by inferring only a missing state; an explicit "unknown"
        # remains excluded from every asserted denominator.
        if not website_state:
            website_state = "present" if expected_hosts else "absent"
        if not email_state:
            email_state = "present" if _emails(expected.get("Expected Email")) else "absent"
        if not phone_state:
            phone_state = "present" if _phones(expected.get("Expected Phone")) else "absent"
        actual = actual_by_company.get(key, {})
        candidate_row = candidates_by_company.get(key, {})
        published_host = _host(actual.get("website"))
        selected_host = _host(candidate_row.get("selected_website"))
        if website_state not in {"present", "absent"}:
            counts["website_unknown_rows"] += 1
            continue
        counts["website_asserted_rows"] += 1
        if not published_host:
            counts["abstained"] += 1
        else:
            counts["published_count"] += 1
            if website_state == "present" and published_host in expected_hosts:
                counts["published_correct"] += 1
        if selected_host:
            counts["selected_count"] += 1
            if website_state == "present" and selected_host in expected_hosts:
                counts["selected_correct"] += 1
        if website_state != "present" or not expected_hosts:
            continue
        counts["expected_websites"] += 1
        candidate_hosts = [
            _host(candidate_row.get(f"candidate_{index}_url")) for index in range(1, 4)
        ]
        if candidate_hosts and candidate_hosts[0] in expected_hosts:
            counts["candidate_top1_hits"] += 1
        if expected_hosts.intersection(host for host in candidate_hosts if host):
            counts["candidate_top3_hits"] += 1
        if published_host in expected_hosts:
            counts["correct_site_rows"] += 1
            extracted_emails = set(_emails(actual.get("email"))) | set(_emails(actual.get("alternative_emails")))
            extracted_phones = set(_phones(actual.get("phone"))) | set(_phones(actual.get("alternative_phones")))
            if email_state == "present":
                counts["correct_site_email_asserted"] += 1
                if extracted_emails & set(_emails(expected.get("Expected Email"))):
                    counts["email_on_correct_site"] += 1
            if phone_state == "present":
                counts["correct_site_phone_asserted"] += 1
                if extracted_phones & set(_phones(expected.get("Expected Phone"))):
                    counts["phone_on_correct_site"] += 1

    def ratio(numerator: int, denominator: int) -> float:
        return round(numerator / denominator, 4) if denominator else 0.0

    counts.update({
        "total_companies": total,
        "candidate_recall_at_1": ratio(counts["candidate_top1_hits"], counts["expected_websites"]),
        "candidate_recall_at_3": ratio(counts["candidate_top3_hits"], counts["expected_websites"]),
        "selection_accuracy": ratio(counts["selected_correct"], counts["selected_count"]),
        "publication_precision": ratio(counts["published_correct"], counts["published_count"]),
        "abstention_rate": ratio(counts["abstained"], counts["website_asserted_rows"]),
        "email_recall_given_correct_site": ratio(
            counts["email_on_correct_site"],
            counts["correct_site_email_asserted"],
        ),
        "phone_recall_given_correct_site": ratio(
            counts["phone_on_correct_site"],
            counts["correct_site_phone_asserted"],
        ),
    })
    return counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Manual golden XLSX ile pipeline sonucunu karsilastirir.")
    parser.add_argument("--expected", type=Path, required=True)
    parser.add_argument("--actual", type=Path, required=True)
    parser.add_argument("--candidates", type=Path)
    args = parser.parse_args()
    metrics, complete_matches = evaluate(args.expected, args.actual)
    for field in FIELDS:
        values = metrics[field]
        denominator = values["tp"] + values["fp"]
        precision = values["tp"] / denominator * 100 if denominator else 0.0
        print(
            f"{field}: TP {values['tp']}, FP {values['fp']}, FN {values['fn']}, "
            f"precision %{precision:.1f}"
        )
    print(f"tam firma eslesmesi: {len(complete_matches)}/{len(_sheet_rows(args.expected, 'Manual Report'))}")
    if args.candidates:
        stages = evaluate_stages(args.expected, args.actual, args.candidates)
        print("--- asama metrikleri ---")
        for key in (
            "candidate_recall_at_1", "candidate_recall_at_3", "selection_accuracy",
            "publication_precision", "abstention_rate",
            "email_recall_given_correct_site", "phone_recall_given_correct_site",
        ):
            print(f"{key}: %{stages[key] * 100:.1f}")


if __name__ == "__main__":
    main()
