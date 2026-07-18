import argparse
from pathlib import Path

from openpyxl import load_workbook

import config
from modules import excel


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Eski supheli satirlar ile yeni retry sonuclarini yan yana koyar.")
    parser.add_argument("--old", type=Path, default=config.OUTPUT_DIR / "suspicious_contacts.xlsx")
    parser.add_argument("--new", type=Path, default=config.OUTPUT_DIR / "brightdata_suspicious_retry" / "contacts.xlsx")
    parser.add_argument("--output", type=Path, default=config.OUTPUT_DIR / "suspicious_old_vs_new.xlsx")
    return parser.parse_args()


def _load(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {header: row[index] if index < len(row) and row[index] is not None else "" for index, header in enumerate(headers)}
        for row in rows[1:]
    ]


def _key(row: dict) -> str:
    return str(row.get("company", "")).strip().casefold()


def main() -> None:
    args = parse_args()
    old_rows = _load(args.old)
    new_by_key = {_key(row): row for row in _load(args.new)}
    rows = []
    for old in old_rows:
        new = new_by_key.get(_key(old), {})
        rows.append(
            {
                "company": old.get("company", ""),
                "old_website": old.get("website", ""),
                "old_email": old.get("email", ""),
                "old_phone": old.get("phone", ""),
                "old_status": old.get("status", ""),
                "old_score": old.get("score", ""),
                "audit_reason": old.get("audit_reason", ""),
                "new_website": new.get("website", ""),
                "new_email": new.get("email", ""),
                "new_phone": new.get("phone", ""),
                "new_status": new.get("status", ""),
                "new_score": new.get("score", ""),
                "new_reason": new.get("reason", ""),
                "decision": "",
            }
        )

    excel._write_rows(
        args.output,
        [
            "company",
            "old_website",
            "old_email",
            "old_phone",
            "old_status",
            "old_score",
            "audit_reason",
            "new_website",
            "new_email",
            "new_phone",
            "new_status",
            "new_score",
            "new_reason",
            "decision",
        ],
        rows,
    )
    print(f"Karsilastirma satiri: {len(rows)}")
    print(f"Yazildi: {args.output}")


if __name__ == "__main__":
    main()
